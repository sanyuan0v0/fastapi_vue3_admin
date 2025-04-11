# -*- coding: utf-8 -*-

import io
import pandas as pd
from typing import Any, Dict, List
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


class ExcelUtil:
    """Excel文件处理工具类"""
    
    @classmethod
    def __mapping_list(cls, list_data: List[Dict[str, Any]], mapping_dict: Dict) -> List:
        """
        工具方法：将list数据中的字段名映射为对应的中文字段名

        :param list_data: 数据列表
        :param mapping_dict: 映射字典
        :return: 映射后的数据列表
        """
        mapping_data = [{mapping_dict.get(key): item.get(key) for key in mapping_dict} for item in list_data]

        return mapping_data
    
    @classmethod
    def get_excel_template(cls, header_list: List[str], selector_header_list: List[str], option_list: List[Dict[str, List[str]]]) -> bytes:
        """
        生成Excel模板文件

        Args:
            header_list: 表头列表
            selector_header_list: 需要设置下拉选择的表头列表 
            option_list: 下拉选项配置列表

        Returns:
            bytes: Excel文件的二进制数据
        """
        wb = Workbook()
        ws = wb.active

        # 设置表头样式
        header_fill = PatternFill(start_color='ababab', end_color='ababab', fill_type='solid')
        center_align = Alignment(horizontal='center')

        # 写入表头
        for col_num, header in enumerate(header_list, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.alignment = center_align
            ws.column_dimensions[get_column_letter(col_num)].width = 12

        # 设置下拉选择
        for selector_header in selector_header_list:
            col_idx = header_list.index(selector_header) + 1
            
            # 获取当前表头的选项列表
            header_options = next((opt.get(selector_header) for opt in option_list if selector_header in opt), [])
            
            if header_options:
                dv = DataValidation(type='list', formula1=f'"{",".join(header_options)}"')
                dv.add(f'{get_column_letter(col_idx)}2:{get_column_letter(col_idx)}1048576')
                ws.add_data_validation(dv)

        # 导出为二进制数据
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    @classmethod
    def export_list2excel(cls, list_data: List[Dict[str, Any]], mapping_dict: Dict) -> bytes:
        """
        将列表数据导出为Excel

        Args:
            list_data: 要导出的数据列表

        Returns:
            bytes: Excel文件的二进制数据
        """
        mapping_data = cls.__mapping_list(list_data, mapping_dict)
        df = pd.DataFrame(mapping_data)
        buffer = io.BytesIO()
        df.to_excel(buffer, index=False, engine='openpyxl')
        return buffer.getvalue()
